import openpyxl
from datetime import datetime, date

#Reading xlsx file
wb = openpyxl.load_workbook(filename = 'balances.xlsx')
ws = wb.active

print(datetime.today().month)
print(int(wb.sheetnames[-1]))

if int(datetime.today().month) != int(wb.sheetnames[-1]):

	min_row = ws.min_row + 1
	blank_col = ws.min_column

	sheet_new = wb.create_sheet(str(datetime.today().month))
	wb.save('balances.xlsx')
	ws = wb.active
	list_of_names = list('')
	for row in ws.iter_rows(min_row=min_row, max_col=1, max_row=ws.max_row, values_only=True):
		list_of_names.append(str(row[0]))
	print(list_of_names)
	for row in range(len(list_of_names) + 1):
		sheet_new.cell(row=row+1, column=blank_col, value=str(list_of_names[row - 1]))

print(ws)
#Calling particular sheet in xlsx file
sheet = wb[str(datetime.today().month)]

#Finding first blank column
blank_col = sheet.max_column + 1

min_row = sheet.min_row + 1

#Getting data from the user
data_set = list('')
for row in ws.iter_rows(min_row=min_row, max_col=1, max_row=sheet.max_row, values_only=True):
	data_set.append(input(str(row[0]) + ": "))
print(data_set)

#Inserting new data in the blank column
for row in range(1,len(data_set) + 1):
	sheet.cell(row=row+1, column=blank_col, value=int(data_set[row - 1]))

#Календарь
date = str(date.today())
print(date)

# Заполнить дату для трат
sheet.cell(row=1, column=blank_col, value=date)    

#ws.insert_cols(max_col + 1,1)


#sheet.insert_cols(max_col, data_set)  

'''colA = ws['A']
col_range = ws['A:D']
row3 = ws[3]
row_range = ws[5:10]

print(col_range)
print(row_range)
'''

wb.save('balances.xlsx')

# Got GIT